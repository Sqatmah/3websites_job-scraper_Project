import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from datetime import datetime

def style_sheet(ws):
    header_font = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="4F81BD")

    for col_num, column_cells in enumerate(ws.iter_rows(min_row=1, max_row=1), start=1):
        for cell in column_cells:
            cell.font = header_font
            cell.fill = fill
            # Optional: Auto-adjust column width
            max_length = max(len(str(cell.value)) for cell in ws[get_column_letter(col_num)])
            ws.column_dimensions[get_column_letter(col_num)].width = max(15, max_length + 2)

def save_to_excel(jobs: list):
    date = datetime.now().strftime("%Y-%m-%d")
    filename = f"data/jobs_{date}.xlsx"

    # Organize jobs by source
    sources = {}
    for job in jobs:
        source = job.get("Source", "Unknown")
        sources.setdefault(source, []).append(job)

    # Write each source to a separate sheet
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        for source, job_list in sources.items():
            df = pd.DataFrame(job_list)
            sheet_name = source[:31]  # Excel limit
            df.to_excel(writer, index=False, sheet_name=sheet_name)

    # Style each sheet
    wb = load_workbook(filename)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        style_sheet(ws)
    wb.save(filename)

    return filename  # ✅ Return for emailing
